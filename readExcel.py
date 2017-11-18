import openpyxl
import os
# set working directory where excel file is
os.chdir("/Users/hirenkumarpatel/Desktop/pyth/exelPytha")
# assign existing file to variable workbook
workbook = openpyxl.load_workbook('example.xlsx')

print(type(workbook))
#inside of workbook select Sheet1 and assign to variable sheet
sheet = workbook.get_sheet_by_name('Sheet1')

print(type(sheet))

#to get sheet name in workbook

sheetNames = workbook.get_sheet_names()
print(sheetNames)

#to read cell inside sheet

cell = sheet['A1']
print('cell name {} and its value {} '.format(cell,cell.value))

while True:
    yourCell = input("your cell name> ")
    if yourCell == "":
        print("no value added")
        continue
    if yourCell == "done" or yourCell == "DONE":
        break
    else:
        try:
            print(sheet[yourCell].value)
        except:
            print("no proper input")


# ENTERING ROW NUMBER AND COLUMN NUMBER YOU GET OR READ VALUE INSIDE CELL\

getCellValue = str(sheet.cell(row=8,column=2).value)
print(getCellValue)

def findCell():
    search = input("enter your word >")
    for i in range(1, 9):
        cellValue = sheet.cell(row=i, column=2)
        print(cellValue.value)
        print(str(cellValue.value))
        if  search == str(cellValue.value):
            nextCellValue = sheet.cell(row=i, column=3)
            print("your next adjacent cell {} Bingo".format(nextCellValue.value))
            break
        else:
            print("your word not found in the row!!!")
            continue



findCell()

#    print(sheet[yourCell].value)














