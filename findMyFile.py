import glob, os
import openpyxl



def find(name, path):
    #path = "/Users/hirenkumarpatel/Desktop/pyth/*/"
    file_dir_extension = os.path.join(path, name)
    # "/Users/hirenkumarpatel/Desktop/pyth/*/*.xlsx"
    print(file_dir_extension)
    for f in glob.iglob(file_dir_extension):  # generator, search immediate subdirectories
        print(f)

def dailyDownloadFile():
    import time
    import datetime
    mydate = datetime.datetime.now()
    file_extn = "CFS_ANZ_Daily_Download_"
    currentMonth = mydate.strftime("%b")
    print(mydate.strftime("%b"))
    todayDate = time.strftime(currentMonth+"_%dth_%Y.xlsx")
    print(file_extn+todayDate)
    return  file_extn+todayDate

def search(excelPath,file):
    os.chdir(excelPath)
    # assign existing file to variable workbook
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name('Sheet1')

    search = input("enter your word >")
    for i in range(1, 9):
        cellValue = sheet.cell(row=i, column=2)
        print(cellValue.value)
        print(str(cellValue.value))
        if search == str(cellValue.value):
            nextCellValue = sheet.cell(row=i, column=3)
            print("your next adjacent cell {} Bingo".format(nextCellValue.value))
            break
        else:
            print("your word not found in the row!!!")
            continue

while True:
    givenName = input('Give file name OR press Enter to get name of daily download excel file ? >')
    if givenName == "":
        givenName = dailyDownloadFile()

    if givenName == 'done'or givenName == 'exit':
        break
    if givenName == 'help':
        print('type in Give File Nmae *.extention or filr name')
        print('type in Give File Path /dir1/secondary diectory/')
        print("type 'done' or 'exit' to exit programe ")
        print("type 'sifile' to get daily down load xlsx file")
    if givenName == "sifile":
         pathName = input('give path to get Daily down load file ? >')
         givenName = dailyDownloadFile()
         print("pathName is {} file is {}".format(pathName, givenName))
         search(pathName, givenName)

         continue

    else:
        pathName = input('Give file path >')
        pathNameExten = os.path.join(pathName, '*/')
        find(givenName, pathNameExten)
        continue





